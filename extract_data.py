"""Extract quiz data from Excel to JSON."""
import json
import openpyxl
import re

EXCEL_FILE = "毛中特题库（700_400）.xlsx"
OUTPUT_FILE = "questions.json"

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

def safe_str(v):
    """Convert cell value to stripped string, return '' for None."""
    if v is None:
        return ""
    s = str(v).strip()
    return s

def extract_single(ws):
    questions = []
    for r in range(2, ws.max_row + 1):
        question = safe_str(ws.cell(row=r, column=1).value)
        if not question:
            continue
        q = {
            "type": "single",
            "qid": f"S{len(questions) + 1}",
            "id": len(questions) + 1,
            "question": question,
            "options": {
                "A": safe_str(ws.cell(row=r, column=2).value),
                "B": safe_str(ws.cell(row=r, column=3).value),
                "C": safe_str(ws.cell(row=r, column=4).value),
                "D": safe_str(ws.cell(row=r, column=5).value),
            },
            # Col12 = correct answer (Col7 is user answer column, might differ)
            "answer": safe_str(ws.cell(row=r, column=12).value),
        }
        # Validate answer
        if q["answer"] and q["answer"] not in "ABCD":
            # Could be a multi-char answer, just clean it
            q["answer"] = q["answer"].strip().upper()
            # Only keep first valid letter if it's garbled
            valid = [c for c in q["answer"] if c in "ABCD"]
            if valid:
                q["answer"] = valid[0] if len(valid) == 1 else "".join(valid)
        questions.append(q)
    return questions

def extract_multi(ws):
    questions = []
    for r in range(2, ws.max_row + 1):
        question = safe_str(ws.cell(row=r, column=1).value)
        if not question:
            continue
        answer = safe_str(ws.cell(row=r, column=6).value)
        # Validate answer contains only ABCD letters
        answer = "".join(c for c in answer.upper() if c in "ABCD")
        q = {
            "type": "multi",
            "qid": f"M{len(questions) + 1}",
            "id": len(questions) + 1,
            "question": question,
            "options": {
                "A": safe_str(ws.cell(row=r, column=2).value),
                "B": safe_str(ws.cell(row=r, column=3).value),
                "C": safe_str(ws.cell(row=r, column=4).value),
                "D": safe_str(ws.cell(row=r, column=5).value),
            },
            "answer": answer,
        }
        questions.append(q)
    return questions

def extract_judge(ws):
    questions = []
    for r in range(2, ws.max_row + 1):
        question = safe_str(ws.cell(row=r, column=1).value)
        if not question:
            continue
        ans = safe_str(ws.cell(row=r, column=2).value)
        if ans in ("对", "√", "A", "正确", "T", "TRUE", "是"):
            ans = "对"
        elif ans in ("错", "×", "B", "错误", "F", "FALSE", "否"):
            ans = "错"
        # If ans is something else, keep it as-is (shouldn't happen)
        q = {
            "type": "judge",
            "qid": f"J{len(questions) + 1}",
            "id": len(questions) + 1,
            "question": question,
            "answer": ans,
        }
        questions.append(q)
    return questions

print("Extracting...")

single_qs = extract_single(wb["单选题"])
print(f"  单选题: {len(single_qs)} 道")

multi_qs = extract_multi(wb["多选题"])
print(f"  多选题: {len(multi_qs)} 道")

judge_qs = extract_judge(wb["判断题"])
print(f"  判断题: {len(judge_qs)} 道")

all_data = {
    "single": single_qs,
    "multi": multi_qs,
    "judge": judge_qs,
    "counts": {
        "single": len(single_qs),
        "multi": len(multi_qs),
        "judge": len(judge_qs),
        "total": len(single_qs) + len(multi_qs) + len(judge_qs),
    }
}

with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\nSaved to {OUTPUT_FILE}")
print(f"Total: {all_data['counts']['total']} questions")

# Quick validation
print("\nValidation:")
for key in ["single", "multi", "judge"]:
    empty_q = sum(1 for q in all_data[key] if not q["question"])
    empty_a = sum(1 for q in all_data[key] if not q["answer"])
    print(f"  {key}: {empty_q} empty questions, {empty_a} empty answers")
print("Done.")
